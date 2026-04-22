from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


WORKBOOK_PATH = Path(
    r"Z:\HSEQS\15 - Audit Reports\ENS-HSEQ-SCH-001 HSEQ Audit Schedule 07.01.2026 Rev C.xlsx"
)
OUTPUT_DIR = Path("scripts") / "output" / "audit-import-2026"
AUDIT_SHEET = "Audit Schedule"
FINDING_SHEET = "Audit Findings"
TARGET_YEAR = 2026


AUDIT_REFERENCE_CORRECTIONS = {
    ("C-Job Naval Architects", "SUP-26-001"): "SUP-26-010",
    ("UTEC", "SUP-26-002"): "SUP-26-011",
}

FINDING_REFERENCE_CORRECTIONS = {
    "SUP-26-012": "INT-26-012",
    "SUP-26-016": "INT-26-016",
}

MONTH_MAP = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}


@dataclass
class HeldBackRow:
    sheet: str
    row_number: int
    reason: str
    source: dict[str, Any]

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "row_number": self.row_number,
            "reason": self.reason,
            "source": self.source,
        }


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.upper() == "N/A":
        return None
    return text


def normalize_bool(value: Any) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"true", "yes", "y", "1"}:
        return True
    if text in {"false", "no", "n", "0"}:
        return False
    return None


def normalize_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text or text.upper() == "N/A":
        return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def normalize_month(year_value: Any, month_value: Any) -> str | None:
    year = int(year_value) if year_value is not None else None
    if year != TARGET_YEAR:
        return None

    if month_value is None:
        return None

    if isinstance(month_value, datetime):
        return f"{month_value.year:04d}-{month_value.month:02d}"

    text = str(month_value).strip()
    if not text:
        return None

    match = re.match(r"^(\d{4})-(\d{2})$", text)
    if match:
        return text

    month_num = MONTH_MAP.get(text.lower())
    if month_num:
        return f"{TARGET_YEAR:04d}-{month_num:02d}"

    return None


def normalize_audit_type(value: Any) -> str | None:
    text = normalize_text(value)
    if not text:
        return None

    lowered = text.lower()
    if lowered == "internal":
        return "Internal"
    if lowered == "external":
        return "External"
    if lowered == "supplier":
        return "Supplier"
    return text


def normalize_finding_category(value: Any) -> str | None:
    text = normalize_text(value)
    if not text:
        return None

    lowered = text.lower()
    if "major" in lowered:
        return "Major"
    if "minor" in lowered:
        return "Minor"
    if "ofi" in lowered:
        return "OFI"
    if "obs" in lowered:
        return "OBS"
    return None


def normalize_standards(value: Any) -> list[str]:
    text = normalize_text(value)
    return [text] if text else []


def serialize_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return rows


def ensure_output_dir() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def write_json(name: str, payload: Any) -> None:
    path = OUTPUT_DIR / name
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=True), encoding="utf-8")


def row_to_source(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    source: dict[str, Any] = {}
    for index, header in enumerate(headers):
        value = values[index] if index < len(values) else None
        if isinstance(value, datetime):
            source[header] = value.date().isoformat()
        else:
            source[header] = value
    return source


def load_sheet_rows(sheet_name: str) -> tuple[list[str], list[tuple[int, tuple[Any, ...]]]]:
    wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    ws = wb[sheet_name]
    headers = [str(cell).strip() if cell is not None else f"column_{idx + 1}" for idx, cell in enumerate(next(ws.iter_rows(min_row=7, max_row=7, values_only=True)))]
    rows = [(row_number, row) for row_number, row in enumerate(ws.iter_rows(min_row=8, values_only=True), start=8)]
    wb.close()
    return headers, rows


def stage_audits() -> tuple[list[dict[str, Any]], list[HeldBackRow], dict[str, Any]]:
    headers, rows = load_sheet_rows(AUDIT_SHEET)
    staged: list[dict[str, Any]] = []
    held_back: list[HeldBackRow] = []
    corrections_applied: list[dict[str, Any]] = []

    for row_number, row in rows:
        if not row or row[0] != TARGET_YEAR:
            continue

        source = row_to_source(headers, row)
        function_name = normalize_text(row[3])
        original_reference = normalize_text(row[6])

        if not original_reference:
            held_back.append(
                HeldBackRow(
                    sheet=AUDIT_SHEET,
                    row_number=row_number,
                    reason="Missing Audit Reference",
                    source=source,
                )
            )
            continue

        corrected_reference = AUDIT_REFERENCE_CORRECTIONS.get(
            (function_name or "", original_reference),
            original_reference,
        )
        if corrected_reference != original_reference:
            corrections_applied.append(
                {
                    "sheet": AUDIT_SHEET,
                    "row_number": row_number,
                    "function": function_name,
                    "from": original_reference,
                    "to": corrected_reference,
                }
            )

        completed = normalize_bool(row[14])
        status = "Completed" if completed else "Planned"

        staged.append(
            {
                "sheet": AUDIT_SHEET,
                "source_row_number": row_number,
                "source_year": row[0],
                "original_audit_reference": original_reference,
                "audit_number": corrected_reference,
                "title": function_name,
                "audit_type": normalize_audit_type(row[4]),
                "auditee": function_name,
                "lead_auditor": normalize_text(row[8]),
                "audit_date": normalize_date(row[9]),
                "audit_month": normalize_month(row[0], row[1]),
                "status": status,
                "standards": normalize_standards(row[7]),
                "procedure_reference": None,
                "certification_body": None,
                "location": normalize_text(row[5]),
                "linked_ncrs": [],
                "linked_actions": [],
                "source_snapshot": source,
            }
        )

    duplicate_refs = {
        ref: count
        for ref, count in Counter(item["audit_number"] for item in staged).items()
        if count > 1
    }

    report = {
        "sheet": AUDIT_SHEET,
        "in_scope_rows": len(staged) + len(held_back),
        "staged_rows": len(staged),
        "held_back_rows": len(held_back),
        "manual_reference_corrections_applied": corrections_applied,
        "duplicate_audit_numbers_after_corrections": duplicate_refs,
    }
    return staged, held_back, report


def stage_findings(valid_audit_numbers: set[str]) -> tuple[list[dict[str, Any]], list[HeldBackRow], dict[str, Any]]:
    headers, rows = load_sheet_rows(FINDING_SHEET)
    staged: list[dict[str, Any]] = []
    held_back: list[HeldBackRow] = []
    per_audit_sequence: defaultdict[str, int] = defaultdict(int)
    corrections_applied: list[dict[str, Any]] = []

    for row_number, row in rows:
        if not row or row[0] != TARGET_YEAR:
            continue

        source = row_to_source(headers, row)
        original_reference = normalize_text(row[2])
        if not original_reference:
            held_back.append(
                HeldBackRow(
                    sheet=FINDING_SHEET,
                    row_number=row_number,
                    reason="Missing Audit Reference",
                    source=source,
                )
            )
            continue

        corrected_reference = FINDING_REFERENCE_CORRECTIONS.get(original_reference, original_reference)
        if corrected_reference != original_reference:
            corrections_applied.append(
                {
                    "sheet": FINDING_SHEET,
                    "row_number": row_number,
                    "from": original_reference,
                    "to": corrected_reference,
                }
            )

        if corrected_reference not in valid_audit_numbers:
            held_back.append(
                HeldBackRow(
                    sheet=FINDING_SHEET,
                    row_number=row_number,
                    reason=f"No staged parent audit for reference {corrected_reference}",
                    source=source,
                )
            )
            continue

        category = normalize_finding_category(row[5])
        if not category:
            held_back.append(
                HeldBackRow(
                    sheet=FINDING_SHEET,
                    row_number=row_number,
                    reason="Unmapped finding category",
                    source=source,
                )
            )
            continue

        per_audit_sequence[corrected_reference] += 1
        finding_reference = f"{corrected_reference}-F{per_audit_sequence[corrected_reference]:02d}"

        completed = normalize_bool(row[15])
        status = "Closed" if completed else "Open"

        staged.append(
            {
                "sheet": FINDING_SHEET,
                "source_row_number": row_number,
                "source_year": row[0],
                "original_audit_reference": original_reference,
                "audit_reference": corrected_reference,
                "reference": finding_reference,
                "clause": None,
                "category": category,
                "description": normalize_text(row[9]),
                "owner": normalize_text(row[8]),
                "status": status,
                "due_date": normalize_date(row[13]),
                "closure_date": normalize_date(row[14]),
                "root_cause": normalize_text(row[11]),
                "containment_action": normalize_text(row[10]),
                "corrective_action": normalize_text(row[12]),
                "source_snapshot": source,
            }
        )

    unresolved_parent_refs = sorted(
        {
            item.source.get("Audit Reference")
            for item in held_back
            if item.reason.startswith("No staged parent audit")
        }
    )

    report = {
        "sheet": FINDING_SHEET,
        "in_scope_rows": len(staged) + len(held_back),
        "staged_rows": len(staged),
        "held_back_rows": len(held_back),
        "manual_reference_corrections_applied": corrections_applied,
        "unresolved_parent_references_after_corrections": unresolved_parent_refs,
    }
    return staged, held_back, report


def build_conflict_report(
    staged_audits: list[dict[str, Any]],
    staged_findings: list[dict[str, Any]],
    held_back: list[HeldBackRow],
    audit_report: dict[str, Any],
    finding_report: dict[str, Any],
) -> dict[str, Any]:
    parent_refs = {row["audit_number"] for row in staged_audits}
    child_refs = {row["audit_reference"] for row in staged_findings}

    audits_without_findings = sorted(parent_refs - child_refs)
    held_back_by_reason = Counter(item.reason for item in held_back)

    return {
        "workbook_path": str(WORKBOOK_PATH),
        "target_year": TARGET_YEAR,
        "audit_report": audit_report,
        "finding_report": finding_report,
        "summary": {
            "staged_audits": len(staged_audits),
            "staged_findings": len(staged_findings),
            "held_back_rows": len(held_back),
            "audits_without_findings": audits_without_findings,
            "held_back_by_reason": dict(held_back_by_reason),
        },
    }


def main() -> None:
    ensure_output_dir()

    staged_audits, held_back_audits, audit_report = stage_audits()
    valid_audit_numbers = {row["audit_number"] for row in staged_audits}
    staged_findings, held_back_findings, finding_report = stage_findings(valid_audit_numbers)

    held_back_rows = [item.to_dict() for item in [*held_back_audits, *held_back_findings]]
    conflict_report = build_conflict_report(
        staged_audits,
        staged_findings,
        [*held_back_audits, *held_back_findings],
        audit_report,
        finding_report,
    )

    import_rows = {
        "audits": staged_audits,
        "audit_findings": staged_findings,
    }

    write_json("normalized_audits_2026.json", serialize_rows(staged_audits))
    write_json("normalized_findings_2026.json", serialize_rows(staged_findings))
    write_json("conflict_report_2026.json", conflict_report)
    write_json("import_rows_2026.json", import_rows)
    write_json("held_back_rows_2026.json", held_back_rows)

    print(f"Staged audits: {len(staged_audits)}")
    print(f"Staged findings: {len(staged_findings)}")
    print(f"Held back rows: {len(held_back_rows)}")
    print(f"Output directory: {OUTPUT_DIR.resolve()}")


if __name__ == "__main__":
    main()
